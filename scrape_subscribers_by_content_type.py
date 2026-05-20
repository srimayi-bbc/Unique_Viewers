"""
Subscribers Gained by Content Type — Scraper
=============================================
Reads YouTube Studio analytics URLs from YouTube-manual-WS-New_Users.csv,
opens each page in parallel browser tabs, and extracts the
"Subscribers Gained" volume + percentage for each content-type row
(Videos, Shorts, Live streams).

Output Excel columns
--------------------
    Channel | Channel_title
    | videos_volume    | videos_percentage
    | shorts_volume    | shorts_percentage
    | live_stream_volume | live_stream_percentage

Usage
-----
    python scrape_subscribers_by_content_type.py
    python scrape_subscribers_by_content_type.py --tabs 6
    python scrape_subscribers_by_content_type.py --output my_results.xlsx
"""

import argparse
import logging
import os
import time
from pathlib import Path
from urllib.parse import urlparse

# CWE-532 guard: strip Playwright debug env-vars that can leak credentials
for _var in ("DEBUG", "PLAYWRIGHT_DEBUG", "PWDEBUG"):
    os.environ.pop(_var, None)

import pandas as pd
from playwright.sync_api import Page, TimeoutError as PWTimeout, sync_playwright

ALLOWED_HOSTS = {"studio.youtube.com"}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# File paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
WS_CSV = BASE_DIR / "YouTube-manual-WS-New_Users.csv"
DEFAULT_OUTPUT_XLSX = BASE_DIR / "subscribers_by_content_type.xlsx"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_volume(text: str) -> int | None:
    """Convert a YouTube-style metric string to an integer.

    Examples:
        '1.23M' -> 1_230_000
        '45.6K' -> 45_600
        '9,812' -> 9_812
        '-1,200' -> -1_200
    """
    text = text.strip().replace(",", "").replace(" ", "")
    if not text or text == "—" or text == "-":
        return None
    multiplier = 1
    negative = text.startswith("-")
    if negative:
        text = text[1:]
    if text.upper().endswith("M"):
        multiplier = 1_000_000
        text = text[:-1]
    elif text.upper().endswith("K"):
        multiplier = 1_000
        text = text[:-1]
    try:
        result = int(float(text) * multiplier)
        return -result if negative else result
    except ValueError:
        return None



def load_ws_csv(path: Path) -> pd.DataFrame:
    """Load the New Users WS CSV and return rows with valid URLs."""
    df = pd.read_csv(path, header=0, names=["Channel", "Channel_title", "URL"])
    df = df[
        df["Channel"].notna()
        & ~df["Channel"].isin(["Channel", "Studios", "Total"])
        & df["URL"].notna()
        & df["URL"].str.startswith("http", na=False)
    ].reset_index(drop=True)
    logger.info("Loaded %d channel URLs from %s", len(df), path.name)
    return df


def _validate_url(url: str) -> bool:
    """Return True only for https URLs on the allowed YouTube Studio host."""
    try:
        parsed = urlparse(url)
        return parsed.scheme == "https" and parsed.netloc in ALLOWED_HOSTS
    except Exception:
        return False


# ---------------------------------------------------------------------------
# JavaScript: dump raw HTML of every metric cell in the first data row
# (used only with --debug-dom to find the real class names)
# ---------------------------------------------------------------------------
DEBUG_DOM_JS = r"""() => {
    const headers = document.querySelectorAll(
        'yta-explore-table-header-cell .debug-metric-title'
    );
    let colIdx = -1;
    const headerInfo = [];
    for (let i = 0; i < headers.length; i++) {
        const t = headers[i].textContent.trim();
        headerInfo.push(`[${i}] ${t}`);
        if (t.toLowerCase().includes('subscriber') && t.toLowerCase().includes('gain')) {
            colIdx = i;
        }
    }

    const rows = document.querySelectorAll('yta-explore-table-row');
    const rowDumps = [];
    for (const row of rows) {
        const labelEl = row.querySelector(
            '.debug-dimension-value, .dimension-value, yta-explore-table-dimension-cell'
        );
        const label = labelEl ? labelEl.textContent.trim() : '(no label)';
        const cells = row.querySelectorAll('.debug-metric-value');
        const cellDump = colIdx >= 0 && colIdx < cells.length
            ? cells[colIdx].outerHTML
            : `(colIdx=${colIdx}, cells=${cells.length})`;
        rowDumps.push(`ROW: ${label}\n${cellDump}`);
    }

    return {
        headers: headerInfo,
        subscribersColIdx: colIdx,
        rows: rowDumps,
    };
}"""


# ---------------------------------------------------------------------------
# JavaScript: extract Subscribers Gained volume for Total + each content type.
# Percentage is computed in Python as (content_type / total) * 100.
# The DOM confirmed each cell only contains the volume number — no percentage
# element exists in the rendered HTML.
# ---------------------------------------------------------------------------
EXTRACT_JS = r"""() => {
    // ---- 1. Find the Subscribers gained column index ----------------------
    const headers = document.querySelectorAll(
        'yta-explore-table-header-cell .debug-metric-title'
    );
    let colIdx = -1;
    for (let i = 0; i < headers.length; i++) {
        const t = headers[i].textContent.trim().toLowerCase();
        if (t.includes('subscriber') && t.includes('gain')) {
            colIdx = i;
            break;
        }
    }
    if (colIdx === -1) return null;

    // ---- 2. Walk all rows and read the volume from the correct cell -------
    const rows = document.querySelectorAll('yta-explore-table-row');
    if (!rows.length) return null;

    const result = {
        total:       null,
        videos:      null,
        shorts:      null,
        live_stream: null,
    };

    for (const row of rows) {
        const labelEl = row.querySelector(
            '.debug-dimension-value, .dimension-value, yta-explore-table-dimension-cell'
        );
        if (!labelEl) continue;
        const label = labelEl.textContent.trim().toLowerCase();

        let key = null;
        if (label === 'total')                                                          key = 'total';
        else if (label.includes('video') && !label.includes('short') && !label.includes('live')) key = 'videos';
        else if (label.includes('short'))                                               key = 'shorts';
        else if (label.includes('live'))                                                key = 'live_stream';
        if (!key) continue;

        const cells = row.querySelectorAll('.debug-metric-value');
        if (colIdx >= cells.length) continue;
        result[key] = cells[colIdx].textContent.trim() || null;
    }

    return result;
}"""


# ---------------------------------------------------------------------------
# Core scrape function
# ---------------------------------------------------------------------------

def scrape_channel(
    page: Page,
    url: str,
    channel_id: str,
    channel_title: str,
) -> dict:
    """Navigate to a channel's analytics page and return extracted metrics."""

    empty = {
        "Channel": channel_id,
        "Channel_title": channel_title,
        "videos_volume": None,
        "videos_percentage": None,
        "shorts_volume": None,
        "shorts_percentage": None,
        "live_stream_volume": None,
        "live_stream_percentage": None,
    }

    if not _validate_url(url):
        logger.error("  Blocked unsafe URL for '%s'", channel_title)
        return empty

    logger.info("-> %-35s  %s", channel_title, channel_id)

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=60_000)
    except Exception as exc:
        logger.error("  Failed to load page for '%s': %s", channel_title, type(exc).__name__)
        return empty

    # Wait for at least one metric value to appear
    try:
        page.wait_for_selector(".debug-metric-value", timeout=25_000)
    except PWTimeout:
        logger.warning("  Table did not render for '%s'", channel_title)
        return empty

    # Allow JS rendering to settle
    page.wait_for_timeout(1_500)

    data = page.evaluate(EXTRACT_JS)

    if not data:
        logger.warning("  FAIL  No data returned for '%s'", channel_title)
        return empty

    def _vol(key):
        raw = data.get(key)
        return parse_volume(str(raw)) if raw else None

    total_vol   = _vol("total")
    videos_vol  = _vol("videos")
    shorts_vol  = _vol("shorts")
    ls_vol      = _vol("live_stream")

    def _pct(vol):
        """Percentage share of total, rounded to 2 dp. Returns None if either value is missing."""
        if vol is None or not total_vol:
            return None
        return round(vol / total_vol * 100, 2)

    row = {
        "Channel":                channel_id,
        "Channel_title":          channel_title,
        "videos_volume":          videos_vol,
        "videos_percentage":      _pct(videos_vol),
        "shorts_volume":          shorts_vol,
        "shorts_percentage":      _pct(shorts_vol),
        "live_stream_volume":     ls_vol,
        "live_stream_percentage": _pct(ls_vol),
    }

    found = [k for k in ("videos", "shorts", "live_stream") if row[f"{k}_volume"] is not None]
    logger.info(
        "  OK  %-35s  total=%s  found: %s",
        channel_title,
        f"{total_vol:,}" if total_vol else "N/A",
        ", ".join(found) if found else "NONE",
    )
    return row


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Extract Subscribers Gained by content type from YouTube Studio."
    )
    parser.add_argument(
        "--profile", default=None, metavar="DIR",
        help="Browser profile directory (default: ./browser_profile)",
    )
    parser.add_argument(
        "--output", default=str(DEFAULT_OUTPUT_XLSX), metavar="XLSX",
        help="Output Excel path (default: subscribers_by_content_type.xlsx)",
    )
    parser.add_argument(
        "--tabs", type=int, default=6, metavar="N",
        help="Number of parallel browser tabs (default: 6)",
    )
    parser.add_argument(
        "--debug-dom", action="store_true",
        help="Dump raw cell HTML for the first channel to debug_cell_html.txt and exit.",
    )
    args = parser.parse_args()

    output_path = Path(args.output)
    profile_dir = (
        str(Path(args.profile).resolve())
        if args.profile
        else str(BASE_DIR / "browser_profile")
    )
    num_tabs = max(1, args.tabs)

    # ---- Load CSV --------------------------------------------------------
    ws_df = load_ws_csv(WS_CSV)
    if ws_df.empty:
        logger.error("No channel URLs found in %s — aborting.", WS_CSV.name)
        return

    tasks = [
        (str(r["Channel"]).strip(), str(r["Channel_title"]).strip(), str(r["URL"]).strip())
        for _, r in ws_df.iterrows()
    ]

    results: list[dict] = []

    # ---- Browser session -------------------------------------------------
    with sync_playwright() as pw:
        context = pw.chromium.launch_persistent_context(
            profile_dir,
            headless=False,
            no_viewport=True,
            channel="chrome",
            args=[
                "--start-maximized",
                "--disable-blink-features=AutomationControlled",
            ],
            ignore_default_args=["--enable-automation"],
        )

        # One-time login check
        page = context.pages[0] if context.pages else context.new_page()
        page.goto("https://studio.youtube.com", wait_until="domcontentloaded")
        print("\n" + "=" * 60)
        print("  A Chrome window has opened.")
        print("  -> If already logged in, press ENTER immediately.")
        print("  -> Otherwise log in first, then press ENTER.")
        print("  (Login is saved — you won't be asked again.)")
        print("=" * 60)
        input("  Press ENTER when ready >  ")
        page.close()

        # ---- Debug-DOM mode: dump cell HTML for the first channel then exit ----
        if args.debug_dom:
            first_id, first_title, first_url = tasks[0]
            dbg_page = context.new_page()
            logger.info("DEBUG-DOM: loading '%s'", first_title)
            dbg_page.goto(first_url, wait_until="domcontentloaded", timeout=60_000)
            try:
                dbg_page.wait_for_selector(".debug-metric-value", timeout=25_000)
            except PWTimeout:
                logger.warning("  Table did not render — try again or increase timeout.")
            dbg_page.wait_for_timeout(2_000)
            dom_data = dbg_page.evaluate(DEBUG_DOM_JS)
            dbg_page.close()
            context.close()

            dump_path = BASE_DIR / "debug_cell_html.txt"
            with open(dump_path, "w", encoding="utf-8") as fh:
                fh.write(f"Channel : {first_title} ({first_id})\n")
                fh.write(f"URL     : {first_url}\n\n")
                fh.write("=== HEADERS ===\n")
                for h in (dom_data or {}).get("headers", []):
                    fh.write(f"  {h}\n")
                fh.write(f"\nSubscribers-gained col index: "
                         f"{(dom_data or {}).get('subscribersColIdx', 'NOT FOUND')}\n\n")
                fh.write("=== ROW CELL HTML ===\n")
                for r in (dom_data or {}).get("rows", []):
                    fh.write(r + "\n" + "-" * 80 + "\n")

            print(f"\n📄  DOM dump saved → {dump_path}")
            print("    Share this file so the exact selectors can be identified.")
            return

        # Round-robin across N parallel tabs
        logger.info("Scraping %d channels using %d tab(s)…", len(tasks), num_tabs)
        start_time = time.time()

        chunks = [tasks[i::num_tabs] for i in range(num_tabs)]
        pages = [context.new_page() for _ in range(num_tabs)]

        try:
            max_len = max(len(c) for c in chunks)
            for step in range(max_len):
                for pg, chunk in zip(pages, chunks):
                    if step >= len(chunk):
                        continue
                    channel_id, channel_title, url = chunk[step]
                    row = scrape_channel(pg, url, channel_id, channel_title)
                    results.append(row)
        finally:
            for pg in pages:
                pg.close()
            context.close()

    elapsed = time.time() - start_time
    extracted = sum(
        1 for r in results
        if any(r.get(k) is not None for k in (
            "videos_volume", "shorts_volume", "live_stream_volume"
        ))
    )
    logger.info(
        "Scraping complete — %d / %d channels had data  (%.0fs elapsed)",
        extracted, len(tasks), elapsed,
    )

    if not results:
        logger.error("No data extracted. Output file was NOT created.")
        return

    # ---- Build DataFrame -------------------------------------------------
    cols = [
        "Channel", "Channel_title",
        "videos_volume", "videos_percentage",
        "shorts_volume", "shorts_percentage",
        "live_stream_volume", "live_stream_percentage",
    ]
    output_df = pd.DataFrame(results, columns=cols)

    # ---- Write Excel with light formatting --------------------------------
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name="Subscribers by Content Type")

        ws = writer.sheets["Subscribers by Content Type"]

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

        # Bold header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Centre-align all data cells; right-align numeric columns
        numeric_cols = {3, 4, 5, 6, 7, 8}  # 1-indexed
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in numeric_cols:
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

    failed = len(output_df) - extracted
    logger.info(
        "Done!  %d row(s) written to %s  (%d with no data)",
        len(output_df), output_path, failed,
    )
    print(f"\n✅  Excel saved → {output_path}")


if __name__ == "__main__":
    main()
