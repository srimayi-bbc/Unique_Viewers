"""
YouTube Studio Gender/Engaged Views Scraper
===========================================
Reads YouTube Studio analytics URLs from YouTube-manual-WS-Gender.csv,
opens each page in parallel browser tabs, and extracts for each channel:

- Engaged views by gender (Female, Male, User-specified)

Output Excel columns:
    Channel | Channel_title |
    female_engaged_views | male_engaged_views | user_specified_engaged_views

Usage:
    python scrape_gender_engaged_views.py
    python scrape_gender_engaged_views.py --tabs 6
    python scrape_gender_engaged_views.py --output my_results.xlsx
"""

import argparse
import logging
import os
import re
import time
from pathlib import Path
from urllib.parse import urlparse

for _var in ("DEBUG", "PLAYWRIGHT_DEBUG", "PWDEBUG"):
    os.environ.pop(_var, None)

import pandas as pd
from playwright.sync_api import Page, TimeoutError as PWTimeout, sync_playwright

ALLOWED_HOSTS = {"studio.youtube.com"}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
WS_CSV = BASE_DIR / "YouTube-manual-WS-Gender.csv"
DEFAULT_OUTPUT_XLSX = BASE_DIR / "gender_engaged_views.xlsx"

# ----------------------
# Helpers
# ----------------------
def parse_metric(text: str) -> float | None:
    # Cells can contain value + share, e.g. "1,234\n52.3%".
    # Keep only the first metric-like token.
    text = text.strip()
    if not text:
        return None

    first_line = text.splitlines()[0].strip()
    m = re.search(r"[-−]?\s*\d[\d,]*(?:\.\d+)?\s*[KkMm]?", first_line)
    if not m:
        return None

    text = m.group(0).replace(" ", "").replace(",", "").replace("−", "-")
    if text in ("—", "-"):
        return None

    multiplier = 1
    negative = text.startswith("-")
    if negative:
        text = text[1:]
    if text.upper().endswith("M"):
        multiplier = 1_000_000
        text = text[:-1]
    elif text.upper().endswith("K"):
        multiplier = 1_000
        text = text[:-1]
    try:
        result = float(text) * multiplier
        return -result if negative else result
    except ValueError:
        return None

def load_ws_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, header=0, names=["Channel", "Channel_title", "URL"])
    df = df[
        df["Channel"].notna()
        & ~df["Channel"].isin(["Channel", "Studios", "Total"])
        & df["URL"].notna()
        & df["URL"].str.startswith("http", na=False)
    ].reset_index(drop=True)
    logger.info("Loaded %d channel URLs from %s", len(df), path.name)
    return df

def _validate_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
        return parsed.scheme == "https" and parsed.netloc in ALLOWED_HOSTS
    except Exception:
        return False

# ----------------------
# JavaScript extractors
# ----------------------
# Debug: capture full HTML of all rows
DEBUG_DOM_JS = r"""() => {
    const headers = document.querySelectorAll('yta-explore-table-header-cell .debug-metric-title');
    const headerInfo = [];
    for (let i = 0; i < headers.length; i++) {
        headerInfo.push(`[${i}] ${headers[i].textContent.trim()}`);
    }
    const rows = document.querySelectorAll('yta-explore-table-row');
    const rowDumps = [];
    for (const row of rows) {
        const labelEl = row.querySelector('.debug-dimension-value, .dimension-value, yta-explore-table-dimension-cell');
        const label = labelEl ? labelEl.textContent.trim() : '(no label)';
        const cells = row.querySelectorAll('.debug-metric-value');
        const cellTexts = Array.from(cells).map(c => c.textContent.trim());
        rowDumps.push(`ROW: ${label}\nCells: ${cellTexts.join(' | ')}`);
    }
    return {headers: headerInfo, rows: rowDumps};
}"""

# 1. Engaged views by gender (from gender breakdown table)
GENDER_JS = r"""() => {
    // Find the Engaged views column index first
    const headers = document.querySelectorAll('yta-explore-table-header-cell .debug-metric-title');
    let colIdx = -1;
    for (let i = 0; i < headers.length; i++) {
        const t = headers[i].textContent.trim().toLowerCase();
        if (t.includes('engaged view')) {
            colIdx = i;
            break;
        }
    }
    if (colIdx === -1) {
        // Fallback: use first metric column if header text differs by locale/UI variant
        const firstRow = document.querySelector('yta-explore-table-row');
        const firstCells = firstRow ? firstRow.querySelectorAll('.debug-metric-value') : [];
        colIdx = firstCells.length ? 0 : -1;
    }
    if (colIdx === -1) return {female: null, male: null, user: null};

    // Read rows by gender label and fetch the metric from engaged-views column
    const rows = Array.from(document.querySelectorAll('yta-explore-table-row'));
    const result = {female: null, male: null, user: null};
    for (const row of rows) {
        const labelEl = row.querySelector('.debug-dimension-value, .dimension-value, yta-explore-table-dimension-cell');
        if (!labelEl) continue;
        const label = labelEl.textContent.trim().toLowerCase().replace(/\s+/g, ' ');
        const cells = row.querySelectorAll('.debug-metric-value');
        const value = colIdx < cells.length ? ((cells[colIdx].innerText || cells[colIdx].textContent || '').trim() || null) : null;

        // Avoid matching "male" inside "female" by using word boundary for male.
        if (label.includes('female')) result.female = value;
        if (/\bmale\b/.test(label)) result.male = value;
        if (label.includes('user')) result.user = value;
    }
    return result;
}"""

# ----------------------
# Scrape function
# ----------------------
def scrape_channel(page: Page, url: str, channel_id: str, channel_title: str) -> dict:
    empty = {
        "Channel": channel_id,
        "Channel_title": channel_title,
        "female_engaged_views": None,
        "male_engaged_views": None,
        "user_specified_engaged_views": None,
    }
    if not _validate_url(url):
        logger.error("  Blocked unsafe URL for '%s'", channel_title)
        return empty
    logger.info("-> %-35s  %s", channel_title, channel_id)
    try:
        page.goto(url, wait_until="networkidle", timeout=90_000)
    except Exception as exc:
        logger.error("  Failed to load page for '%s': %s", channel_title, type(exc).__name__)
        return empty
    try:
        page.wait_for_selector(".debug-metric-value", timeout=45_000)
    except PWTimeout:
        logger.warning("  Table did not render for '%s'", channel_title)
        return empty
    page.wait_for_timeout(1_500)
    # Engaged views by gender
    gender_data = page.evaluate(GENDER_JS)

    def _vol(raw):
        return parse_metric(str(raw)) if raw else None

    row = {
        "Channel": channel_id,
        "Channel_title": channel_title,
        "female_engaged_views": _vol(gender_data.get("female")) if gender_data else None,
        "male_engaged_views": _vol(gender_data.get("male")) if gender_data else None,
        "user_specified_engaged_views": _vol(gender_data.get("user")) if gender_data else None,
    }
    logger.info(
        "  OK  %-35s  found: %s",
        channel_title,
        ", ".join([k for k in row if row[k] is not None and k.endswith('_engaged_views')])
    )
    return row

# ----------------------
# Main
# ----------------------
def main():
    parser = argparse.ArgumentParser(
        description="Extract Engaged Views by Gender from YouTube Studio."
    )
    parser.add_argument(
        "--profile", default=None, metavar="DIR",
        help="Browser profile directory (default: ./browser_profile)",
    )
    parser.add_argument(
        "--output", default=str(DEFAULT_OUTPUT_XLSX), metavar="XLSX",
        help="Output Excel path (default: gender_engaged_views.xlsx)",
    )
    parser.add_argument(
        "--tabs", type=int, default=4, metavar="N",
        help="Number of parallel browser tabs (default: 4)",
    )
    parser.add_argument(
        "--debug-dom", action="store_true",
        help="Dump raw table HTML for first channel to debug_gender_table.txt and exit.",
    )
    args = parser.parse_args()
    output_path = Path(args.output)
    profile_dir = (
        str(Path(args.profile).resolve())
        if args.profile
        else str(BASE_DIR / "browser_profile")
    )
    num_tabs = max(1, args.tabs)
    ws_df = load_ws_csv(WS_CSV)
    if ws_df.empty:
        logger.error("No channel URLs found in %s — aborting.", WS_CSV.name)
        return
    tasks = [
        (str(r["Channel"]).strip(), str(r["Channel_title"]).strip(), str(r["URL"]).strip())
        for _, r in ws_df.iterrows()
    ]
    results: list[dict] = []
    with sync_playwright() as pw:
        context = pw.chromium.launch_persistent_context(
            profile_dir,
            headless=False,
            no_viewport=True,
            channel="chrome",
            args=[
                "--start-maximized",
                "--disable-blink-features=AutomationControlled",
            ],
            ignore_default_args=["--enable-automation"],
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.goto("https://studio.youtube.com", wait_until="domcontentloaded")
        print("\n" + "=" * 60)
        print("  A Chrome window has opened.")
        print("  -> If already logged in, press ENTER immediately.")
        print("  -> Otherwise log in first, then press ENTER.")
        print("  (Login is saved — you won't be asked again.)")
        print("=" * 60)
        input("  Press ENTER when ready >  ")
        page.close()

        # ---- Debug-DOM mode: dump table HTML for the first channel then exit ----
        if args.debug_dom:
            first_id, first_title, first_url = tasks[0]
            dbg_page = context.new_page()
            logger.info("DEBUG-DOM: loading '%s'", first_title)
            dbg_page.goto(first_url, wait_until="networkidle", timeout=90_000)
            try:
                dbg_page.wait_for_selector(".debug-metric-value", timeout=45_000)
            except PWTimeout:
                logger.warning("  Table did not render — try again or increase timeout.")
            dbg_page.wait_for_timeout(2_000)
            dom_data = dbg_page.evaluate(DEBUG_DOM_JS)
            dbg_page.close()
            context.close()

            dump_path = BASE_DIR / "debug_gender_table.txt"
            with open(dump_path, "w", encoding="utf-8") as fh:
                fh.write(f"Channel: {first_title} ({first_id})\n")
                fh.write(f"URL: {first_url}\n\n")
                fh.write("=== HEADERS ===\n")
                for h in (dom_data or {}).get("headers", []):
                    fh.write(f"  {h}\n")
                fh.write("\n=== ROW DATA ===\n")
                for r in (dom_data or {}).get("rows", []):
                    fh.write(r + "\n" + "-" * 80 + "\n")

            print(f"\n📄  DOM dump saved → {dump_path}")
            print("    Use this to understand the table structure and fix extraction.")
            return

        logger.info("Scraping %d channels using %d tab(s)…", len(tasks), num_tabs)
        start_time = time.time()
        chunks = [tasks[i::num_tabs] for i in range(num_tabs)]
        pages = [context.new_page() for _ in range(num_tabs)]
        try:
            max_len = max(len(c) for c in chunks)
            for step in range(max_len):
                for pg, chunk in zip(pages, chunks):
                    if step >= len(chunk):
                        continue
                    channel_id, channel_title, url = chunk[step]
                    row = scrape_channel(pg, url, channel_id, channel_title)
                    results.append(row)
        finally:
            for pg in pages:
                pg.close()
            context.close()
    elapsed = time.time() - start_time
    logger.info(
        "Scraping complete — %d / %d channels scraped  (%.0fs elapsed)",
        len(results), len(tasks), elapsed,
    )
    if not results:
        logger.error("No data extracted. Output file was NOT created.")
        return
    cols = [
        "Channel", "Channel_title",
        "female_engaged_views", "male_engaged_views", "user_specified_engaged_views",
    ]
    output_df = pd.DataFrame(results, columns=cols)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name="Gender Engaged Views")
        ws = writer.sheets["Gender Engaged Views"]
        for col_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        numeric_cols = {3, 4, 5}  # 1-indexed
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in numeric_cols:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.############"
                else:
                    cell.alignment = Alignment(horizontal="left")
    logger.info(
        "Done!  %d row(s) written to %s",
        len(output_df), output_path,
    )
    print(f"\n✅  Excel saved → {output_path}")

if __name__ == "__main__":
    main()
